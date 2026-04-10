"""
migrate.py — Migrate existing Excel + university_profiles.json into PostgreSQL
==============================================================================
Run ONCE to move all your scraped data into the database.

Usage:
    pip install psycopg2-binary openpyxl
    python migrate.py

Edit DB_CONFIG below to match your PostgreSQL setup.
"""

import json, os, re
from datetime import timezone, datetime
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# ── Database config — edit these ──────────────────────────────
DB_CONFIG = {
    "host":     "localhost",
    "port":     5432,
    "dbname":   "unisearch",
    "user":     "postgres",
    "password": "2000",   # ← change to your PostgreSQL password
}

EXCEL_FILE   = "Uni_data_filled.xlsx"
JSON_FILE    = "university_profiles.json"
NOW          = datetime.now(timezone.utc)

# Section + label mapping for numeric fields
FIELD_META = {
    "acceptance_rate": {
        "section": "admissions",
        "label":   "Acceptance Rate",
        "fmt":     lambda v: f"${int(v):,}" if v else None,
    },
    "in_state_tuition": {
        "section": "tuition",
        "label":   "In-State Tuition",
        "fmt":     lambda v: f"${int(v):,}" if v else None,
    },
    "out_of_state_tuition": {
        "section": "tuition",
        "label":   "Out-of-State Tuition",
        "fmt":     lambda v: f"${int(v):,}" if v else None,
    },
    "total_cost_of_attendance": {
        "section": "tuition",
        "label":   "Total Cost of Attendance",
        "fmt":     lambda v: f"${int(v):,}" if v else None,
    },
    "student_size": {
        "section": "student_life",
        "label":   "Total Enrollment",
        "fmt":     lambda v: f"{int(v):,}" if v else None,
    },
    "average_sat": {
        "section": "admissions",
        "label":   "Average SAT Score",
        "fmt":     lambda v: str(int(v)) if v else None,
    },
    "average_act": {
        "section": "admissions",
        "label":   "Average ACT Score",
        "fmt":     lambda v: str(int(v)) if v else None,
    },
    "graduation_rate_4yr": {
        "section": "outcomes",
        "label":   "4-Year Graduation Rate",
        "fmt":     lambda v: f"{v}%" if v else None,
    },
    "median_earnings_10yr": {
        "section": "outcomes",
        "label":   "Median Earnings (10yr)",
        "fmt":     lambda v: f"${int(v):,}" if v else None,
    },
    "median_debt": {
        "section": "outcomes",
        "label":   "Median Student Debt",
        "fmt":     lambda v: f"${int(v):,}" if v else None,
    },
}


def get_conn():
    return psycopg2.connect(**DB_CONFIG)


def load_excel():
    print(f"Loading {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {headers[i]: row[i] for i in range(len(headers))}
        rows.append(d)
    wb.close()
    print(f"  → {len(rows)} universities in Excel")
    return rows


def load_json():
    if not os.path.exists(JSON_FILE):
        print(f"  WARNING: {JSON_FILE} not found — no rankings/awards/scraped data")
        return {}
    print(f"Loading {JSON_FILE}...")
    with open(JSON_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    print(f"  → {len(data)} entries in JSON")
    return data


def normalize_url(raw):
    if not raw:
        return None
    raw = str(raw).strip()
    if not raw.startswith("http"):
        raw = "https://" + raw
    return raw.rstrip("/")


def run_migration():
    excel_rows = load_excel()
    json_data  = load_json()

    conn = get_conn()
    cur  = conn.cursor()

    # Create schema directly
    print("\nCreating schema...")
    schema_statements = [
        """
        CREATE TABLE IF NOT EXISTS universities (
            id         SERIAL PRIMARY KEY,
            name       TEXT NOT NULL,
            city       TEXT,
            state      TEXT,
            school_url TEXT,
            created_at TIMESTAMPTZ DEFAULT NOW(),
            updated_at TIMESTAMPTZ DEFAULT NOW()
        )
        """,
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_universities_name ON universities(name)",
        "CREATE INDEX IF NOT EXISTS idx_universities_state ON universities(state)",
        """
        CREATE TABLE IF NOT EXISTS university_facts (
            id            SERIAL PRIMARY KEY,
            university_id INTEGER NOT NULL REFERENCES universities(id) ON DELETE CASCADE,
            section       TEXT NOT NULL,
            label         TEXT NOT NULL,
            value         TEXT,
            value_numeric NUMERIC,
            source_url    TEXT,
            extracted_at  TIMESTAMPTZ DEFAULT NOW(),
            extractor     TEXT,
            confidence    NUMERIC(3,2),
            notes         TEXT,
            UNIQUE(university_id, section, label)
        )
        """,
        "CREATE INDEX IF NOT EXISTS idx_facts_university ON university_facts(university_id)",
        "CREATE INDEX IF NOT EXISTS idx_facts_section ON university_facts(section)",
        """
        CREATE TABLE IF NOT EXISTS university_rankings (
            id            SERIAL PRIMARY KEY,
            university_id INTEGER NOT NULL REFERENCES universities(id) ON DELETE CASCADE,
            type          TEXT NOT NULL,
            description   TEXT NOT NULL,
            source_url    TEXT,
            extracted_at  TIMESTAMPTZ DEFAULT NOW(),
            extractor     TEXT
        )
        """,
        "CREATE INDEX IF NOT EXISTS idx_rankings_university ON university_rankings(university_id)",
        """
        CREATE TABLE IF NOT EXISTS carnegie_classifications (
            id                           SERIAL PRIMARY KEY,
            university_id                INTEGER NOT NULL REFERENCES universities(id) ON DELETE CASCADE,
            ipeds_unitid                 TEXT,
            institutional_classification TEXT,
            student_access_earnings      TEXT,
            research_activity            TEXT,
            size_setting                 TEXT,
            basic_classification         TEXT,
            carnegie_page_url            TEXT,
            extracted_at                 TIMESTAMPTZ DEFAULT NOW(),
            confidence                   NUMERIC(3,2),
            match_method                 TEXT,
            notes                        TEXT
        )
        """,
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_carnegie_university ON carnegie_classifications(university_id)",
        """
        CREATE TABLE IF NOT EXISTS scrape_log (
            id            SERIAL PRIMARY KEY,
            university_id INTEGER NOT NULL REFERENCES universities(id) ON DELETE CASCADE,
            scraper       TEXT NOT NULL,
            status        TEXT NOT NULL,
            started_at    TIMESTAMPTZ,
            finished_at   TIMESTAMPTZ,
            facts_found   INTEGER DEFAULT 0,
            error_msg     TEXT
        )
        """,
    ]
    for stmt in schema_statements:
        try:
            cur.execute(stmt)
            conn.commit()
        except Exception as e:
            if "already exists" not in str(e).lower():
                print(f"  Schema warning: {e}")
            conn.rollback()
            cur = conn.cursor()
    print("  Schema ready ✓")

    # Migrate universities
    print(f"\nMigrating {len(excel_rows)} universities...")
    uni_ids = {}
    inserted = 0
    skipped  = 0

    for row in excel_rows:
        name = str(row.get("name") or "").strip()
        if not name:
            continue

        url = normalize_url(row.get("school_url"))

        try:
            cur.execute("""
                INSERT INTO universities (name, city, state, school_url)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (name) DO UPDATE
                    SET city       = EXCLUDED.city,
                        state      = EXCLUDED.state,
                        school_url = EXCLUDED.school_url,
                        updated_at = NOW()
                RETURNING id
            """, (name, row.get("city"), row.get("state"), url))
            uni_id = cur.fetchone()[0]
            uni_ids[name] = uni_id
            inserted += 1
        except Exception as e:
            print(f"  Error inserting {name}: {e}")
            conn.rollback()
            cur = conn.cursor()
            skipped += 1
            continue

    conn.commit()
    print(f"  Universities: {inserted} inserted/updated, {skipped} skipped ✓")

    # Migrate facts from Excel + JSON
    print("\nMigrating facts...")
    facts_inserted = 0
    ranks_inserted = 0

    for row in excel_rows:
        name = str(row.get("name") or "").strip()
        uni_id = uni_ids.get(name)
        if not uni_id:
            continue

        sc = json_data.get(name, {})
        school_url = normalize_url(row.get("school_url"))

        # Numeric fields
        for field, meta in FIELD_META.items():
            # Prefer Excel value, fall back to JSON
            val = row.get(field)
            if val is None:
                val = sc.get(field)
            if val is None:
                continue

            try:
                val_num = float(val)
            except:
                continue

            formatted = meta["fmt"](val_num)
            if not formatted:
                continue

            # Determine source
            source   = "scorecard_api_v1" if sc.get("_source") == "college_scorecard_api" else "web_scraper_v3"
            conf     = 0.90 if source == "scorecard_api_v1" else 0.70
            src_url  = school_url  # will be improved by future scraper

            try:
                cur.execute("""
                    INSERT INTO university_facts
                        (university_id, section, label, value, value_numeric,
                         source_url, extracted_at, extractor, confidence)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (university_id, section, label)
                    DO UPDATE SET
                        value         = EXCLUDED.value,
                        value_numeric = EXCLUDED.value_numeric,
                        source_url    = EXCLUDED.source_url,
                        extracted_at  = EXCLUDED.extracted_at,
                        extractor     = EXCLUDED.extractor,
                        confidence    = EXCLUDED.confidence
                """, (uni_id, meta["section"], meta["label"],
                      formatted, val_num, src_url, NOW, source, conf))
                facts_inserted += 1
            except Exception as e:
                conn.rollback()
                cur = conn.cursor()

        # Rankings & Awards from JSON
        for rtype in ["rankings", "awards"]:
            items = sc.get(rtype)
            if not items:
                continue
            for item in items:
                if not item or not str(item).strip():
                    continue
                try:
                    cur.execute("""
                        INSERT INTO university_rankings
                            (university_id, type, description, source_url,
                             extracted_at, extractor)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (uni_id, rtype[:-1],  # 'ranking' or 'award'
                          str(item).strip(), school_url, NOW, "web_scraper_v3"))
                    ranks_inserted += 1
                except Exception as e:
                    conn.rollback()
                    cur = conn.cursor()

        if facts_inserted % 5000 == 0 and facts_inserted > 0:
            conn.commit()
            print(f"  → {facts_inserted} facts so far...")

    conn.commit()
    print(f"  Facts: {facts_inserted} inserted ✓")
    print(f"  Rankings/Awards: {ranks_inserted} inserted ✓")

    # Summary
    cur.execute("SELECT COUNT(*) FROM universities")
    print(f"\n{'='*50}")
    print(f"MIGRATION COMPLETE")
    print(f"  Universities : {cur.fetchone()[0]:,}")
    cur.execute("SELECT COUNT(*) FROM university_facts")
    print(f"  Facts        : {cur.fetchone()[0]:,}")
    cur.execute("SELECT COUNT(*) FROM university_rankings")
    print(f"  Rankings     : {cur.fetchone()[0]:,}")
    print(f"{'='*50}")

    cur.close()
    conn.close()


if __name__ == "__main__":
    run_migration()
