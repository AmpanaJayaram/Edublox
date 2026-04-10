"""
patch_institution_type.py
=========================
Fills in missing institution_control values in carnegie_classifications
by reading from the Carnegie 2021 Excel file, then falls back to
name-based heuristics for universities without Carnegie data.

Run once: python patch_institution_type.py
"""

import re
import os
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent))
from db_config import get_conn

CARNEGIE_DATA_FILE = "CCIHE2021-PublicData.xlsx"
CONTROL_LABELS = {1: "Public", 2: "Private not-for-profit", 3: "Private for-profit"}

PUBLIC_PATTERNS = [
    r"\bstate university\b", r"\buniversity of \w",
    r"\b\w+ a&m\b", r"\bcommunity college\b",
    r"\btechnical college\b", r"\btechnical university\b",
    r"\bpolytechnic\b", r"\bpublic\b",
]


def infer_control(name: str) -> str | None:
    n = name.lower()
    for pat in PUBLIC_PATTERNS:
        if re.search(pat, n):
            return "Public"
    return None


def main():
    conn = get_conn()
    cur  = conn.cursor()

    cur.execute("""
        ALTER TABLE carnegie_classifications
        ADD COLUMN IF NOT EXISTS institution_control TEXT
    """)
    conn.commit()

    # Step 1: Read from Carnegie Excel if available
    excel_map: dict = {}
    if os.path.exists(CARNEGIE_DATA_FILE):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(CARNEGIE_DATA_FILE, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            uid_col  = headers.index("unitid")  if "unitid"  in headers else None
            ctrl_col = headers.index("control") if "control" in headers else None
            if uid_col is not None and ctrl_col is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    uid  = row[uid_col]
                    ctrl = row[ctrl_col]
                    if uid and ctrl:
                        try:
                            excel_map[str(int(uid))] = CONTROL_LABELS.get(int(ctrl), "")
                        except (ValueError, TypeError):
                            pass
            wb.close()
            print(f"  Loaded {len(excel_map):,} control values from Excel")
        except Exception as e:
            print(f"  Warning: could not read Excel: {e}")
    else:
        print(f"  Note: {CARNEGIE_DATA_FILE} not found — using name heuristics only")

    # Step 2: Update Carnegie records missing institution_control
    cur.execute("""
        SELECT cc.id, cc.university_id, cc.ipeds_unitid,
               u.name, COALESCE(cc.institution_control, '') AS current_control
        FROM carnegie_classifications cc
        JOIN universities u ON u.id = cc.university_id
    """)
    records = cur.fetchall()

    updated_excel     = 0
    updated_heuristic = 0
    already_set       = 0

    for r in records:
        if r["current_control"] and r["current_control"].strip():
            already_set += 1
            continue

        new_ctrl = None
        if r["ipeds_unitid"] and str(r["ipeds_unitid"]) in excel_map:
            new_ctrl = excel_map[str(r["ipeds_unitid"])]
            updated_excel += 1
        else:
            new_ctrl = infer_control(r["name"])
            if new_ctrl:
                updated_heuristic += 1

        if new_ctrl:
            cur.execute(
                "UPDATE carnegie_classifications SET institution_control = %s WHERE id = %s",
                (new_ctrl, r["id"]),
            )

    conn.commit()

    # Step 3: For universities WITHOUT Carnegie data, add institution_type fact
    cur.execute("""
        SELECT u.id, u.name
        FROM universities u
        LEFT JOIN carnegie_classifications cc ON cc.university_id = u.id
        WHERE cc.id IS NULL
    """)
    no_carnegie = cur.fetchall()

    fact_updated = 0
    for r in no_carnegie:
        ctrl = infer_control(r["name"])
        if ctrl:
            cur.execute(
                """
                INSERT INTO university_facts
                    (university_id, section, label, value, extractor, confidence)
                VALUES (%s, 'general', 'Institution Type', %s, 'name_heuristic', 0.70)
                ON CONFLICT (university_id, section, label) DO UPDATE SET value = EXCLUDED.value
                """,
                (r["id"], ctrl),
            )
            fact_updated += 1

    conn.commit()
    conn.close()

    print(f"\n✓ Done!")
    print(f"  Already set:          {already_set:,}")
    print(f"  Updated from Excel:   {updated_excel:,}")
    print(f"  Updated from names:   {updated_heuristic:,}")
    print(f"  Facts added (no-CC):  {fact_updated:,}")
    print(f"\nRestart Flask to see changes.")


if __name__ == "__main__":
    main()
