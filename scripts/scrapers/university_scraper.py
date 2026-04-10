"""
university_scraper.py  (v5 — Hybrid: API for numbers, scraper for rankings/awards)
===================================================================================
WHY HYBRID:
  Most university websites are JavaScript-rendered (React/Vue apps).
  requests+BeautifulSoup cannot see JS-rendered content, so numeric fields
  (tuition, enrollment, SAT etc.) come back empty.

  Solution:
  - College Scorecard API  → all 10 numeric fields (accurate, fast, ~5 min total)
  - Web scraper            → rankings + awards (from the actual university website)

HOW TO GET YOUR FREE API KEY (30 seconds):
  1. Go to: https://api.data.gov/signup
  2. Enter your email — key arrives instantly
  3. Run:  python university_scraper.py YOUR_API_KEY Uni_data.xlsx

WITHOUT API KEY (rankings/awards only):
  python university_scraper.py Uni_data.xlsx

DEBUG A SINGLE UNIVERSITY:
  python university_scraper.py debug https://www.unt.edu/
  python university_scraper.py debug https://www.unt.edu/ YOUR_API_KEY
"""

import os, re, sys, json, time, math, logging
from urllib.parse import urlparse, urljoin
from typing import Optional

import requests
from bs4 import BeautifulSoup
import openpyxl

# ──────────────────────────────────────────────────────
#  CONFIG
# ──────────────────────────────────────────────────────
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"
    )
}
REQUEST_TIMEOUT    = 20
DELAY_PER_REQUEST  = 0.6
SCORECARD_ENDPOINT = "https://api.data.gov/ed/collegescorecard/v1/schools"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s | %(message)s",
    datefmt="%H:%M:%S"
)
log = logging.getLogger(__name__)

FILLABLE_COLS = [
    "acceptance_rate", "in_state_tuition", "out_of_state_tuition",
    "total_cost_of_attendance", "student_size", "average_sat",
    "average_act", "graduation_rate_4yr", "median_earnings_10yr", "median_debt"
]

# ──────────────────────────────────────────────────────
#  COLLEGE SCORECARD API  — fills all 10 numeric fields
#  Data reported directly by universities to US Dept of Education
# ──────────────────────────────────────────────────────
SCORECARD_FIELDS = [
    "school.name",
    "school.school_url",
    "school.ownership",
    "latest.admissions.admission_rate.overall",
    "latest.cost.tuition.in_state",
    "latest.cost.tuition.out_of_state",
    "latest.cost.attendance.academic_year",
    "latest.student.size",
    "latest.admissions.sat_scores.average.overall",
    "latest.admissions.act_scores.midpoint.cumulative",
    "latest.completion.rate_suppressed.four_year",
    "latest.earnings.10_yrs_after_entry.median",
    "latest.aid.median_debt.completers.overall",
]

def fetch_all_scorecard(api_key: str) -> dict:
    """
    Fetches ALL schools from College Scorecard API in one go.
    Returns dict keyed by school name → numeric fields.
    Takes about 3-5 minutes, covers all 6,189 universities.
    """
    log.info("Fetching all universities from College Scorecard API...")

    # Get total count first
    r = requests.get(SCORECARD_ENDPOINT, params={
        "api_key": api_key, "per_page": 1, "page": 0,
        "fields": "school.name"
    }, timeout=30)

    if r.status_code == 403:
        log.error("Invalid API key. Get a free key at https://api.data.gov/signup")
        return {}
    r.raise_for_status()

    total = r.json()["metadata"]["total"]
    pages = math.ceil(total / 100)
    log.info(f"  Total schools: {total:,} — fetching in {pages} pages...")

    all_data = {}
    params = {
        "api_key": api_key,
        "fields": ",".join(SCORECARD_FIELDS),
        "per_page": 100,
    }

    for page in range(pages):
        params["page"] = page
        try:
            resp = requests.get(SCORECARD_ENDPOINT, params=params, timeout=30)
            resp.raise_for_status()
            results = resp.json().get("results", [])
        except Exception as e:
            log.warning(f"  Page {page} failed: {e}, retrying...")
            time.sleep(3)
            try:
                resp = requests.get(SCORECARD_ENDPOINT, params=params, timeout=30)
                results = resp.json().get("results", [])
            except:
                continue

        for s in results:
            name = (s.get("school.name") or "").strip()
            if not name:
                continue

            # acceptance_rate: API returns 0.0–1.0 → store as percentage e.g. 72.3
            acc  = s.get("latest.admissions.admission_rate.overall")
            acc  = round(float(acc) * 100, 1) if acc is not None else None

            # graduation_rate: API returns 0.0–1.0 → store as percentage e.g. 68.0
            grad = s.get("latest.completion.rate_suppressed.four_year")
            grad = round(float(grad) * 100, 1) if grad is not None else None

            earn = s.get("latest.earnings.10_yrs_after_entry.median")
            debt = s.get("latest.aid.median_debt.completers.overall")
            sat  = s.get("latest.admissions.sat_scores.average.overall")
            act  = s.get("latest.admissions.act_scores.midpoint.cumulative")

            all_data[name] = {
                "acceptance_rate":          acc,
                "in_state_tuition":         s.get("latest.cost.tuition.in_state"),
                "out_of_state_tuition":     s.get("latest.cost.tuition.out_of_state"),
                "total_cost_of_attendance": s.get("latest.cost.attendance.academic_year"),
                "student_size":             s.get("latest.student.size"),
                "average_sat":              int(sat)  if sat  else None,
                "average_act":              int(act)  if act  else None,
                "graduation_rate_4yr":      grad,
                "median_earnings_10yr":     int(earn) if earn else None,
                "median_debt":              int(debt) if debt else None,
            }

        fetched = min((page + 1) * 100, total)
        pct = fetched / total * 100
        bar = "█" * int(pct / 4) + "░" * (25 - int(pct / 4))
        print(f"\r  [{bar}] {fetched:,}/{total:,} ({pct:.0f}%)", end="", flush=True)
        time.sleep(0.1)

    print()
    log.info(f"  ✓ Scorecard loaded: {len(all_data):,} schools")
    return all_data


def scorecard_lookup(name: str, url: str, scorecard_data: dict) -> dict:
    """
    Look up a university in the pre-fetched scorecard data.
    Tries exact name match first, then fuzzy domain match.
    """
    # 1. Exact name match
    if name in scorecard_data:
        return scorecard_data[name]

    # 2. Case-insensitive name match
    name_lower = name.lower()
    for k, v in scorecard_data.items():
        if k.lower() == name_lower:
            return v

    # 3. Partial name match (first 20 chars)
    for k, v in scorecard_data.items():
        if name_lower[:20] in k.lower() or k.lower()[:20] in name_lower:
            return v

    return {}


# ──────────────────────────────────────────────────────
#  WEB SCRAPER — rankings & awards only
# ──────────────────────────────────────────────────────
DIRECT_URL_PATTERNS = {
    "rankings": [
        "/rankings", "/rankings/", "/about/rankings",
        "/about/recognition", "/rankings-recognition",
        "/recognition", "/awards", "/about/awards",
        "/accolades", "/about/accolades",
        "/about/rankings-recognition", "/news/rankings",
    ],
    "facts": [
        "/about/facts", "/about/fast-facts", "/about/quick-facts",
        "/quickfacts", "/fast-facts", "/quick-facts",
        "/about/at-a-glance", "/at-a-glance",
        "/allabout", "/allabout/index.html",
        "/about/overview", "/about/profile",
    ],
    "about": [
        "/about", "/about-us", "/about/about-us",
        "/about/index.html",
    ],
}

SUBPAGE_KEYWORDS = {
    "rankings": ["rankings", "recognition", "awards", "accolades", "honors"],
    "facts":    ["fast-facts", "quick-facts", "allabout", "at-a-glance",
                 "facts-figures", "about/facts"],
    "about":    ["/about/", "about-us"],
}

RANKING_PATTERNS = [
    r"#\s*\d+\s+[A-Za-z][^.!?\n]{5,80}",
    r"No\.\s*\d+\s+[A-Za-z][^.!?\n]{5,80}",
    r"ranked\s+#?\s*\d+[^.!?\n]{5,80}",
    r"ranked\s+(?:among|as|one)[^.!?\n]{5,80}",
    r"top\s+\d+\s+[^.!?\n]{5,60}",
    r"top\s+\d+%\s+[^.!?\n]{5,60}",
    r"named\s+(?:a\s+)?(?:top|best|#)[^.!?\n]{5,80}",
    r"best\s+\w+\s+(?:university|college|school|program|value)[^.!?\n]{0,60}",
    r"\d+(?:st|nd|rd|th)\s+(?:in|among|nationally)[^.!?\n]{5,60}",
]

RANKING_BOOST_WORDS = [
    "u.s. news", "us news", "forbes", "princeton review", "money magazine",
    "washington monthly", "times higher", "qs world", "niche",
    "best college", "best university", "national rank", "world rank",
    "carnegie", "r1", "research university", "flagship"
]

AWARD_PATTERNS = [
    r"(?:received?|awarded?|earned?|won)\s+(?:the\s+)?[A-Z][^.!?\n]{5,100}(?:award|prize|grant|recognition|honor|medal|distinction)",
    r"(?:award|prize|recognition|honor|accreditation|certification)\s+(?:from|by|of)\s+[A-Z][^.!?\n]{5,80}",
    r"accredited\s+(?:by|through)\s+[A-Z][^.!?\n]{5,80}",
    r"(?:carnegie|fulbright|nsf|nih|r1|r2)\s+[^.!?\n]{5,80}",
    r"designated\s+(?:a\s+)?[A-Z][^.!?\n]{5,80}",
    r"certified\s+(?:as\s+)?(?:a\s+)?[A-Z][^.!?\n]{5,80}",
]

AWARD_BOOST_WORDS = [
    "award", "prize", "recognition", "honor", "accredit", "certif",
    "designat", "fulbright", "nsf grant", "nih", "excellence",
    "achievement", "distinction", "commend", "merit"
]

NAV_SKIP = ["click here", "learn more", "read more", "apply now",
            "visit us", "contact", "login", "sign in", "search"]


def clean(s) -> str:
    return re.sub(r"\s+", " ", (str(s) if s else "")).strip()

def normalize_url(raw: str) -> str:
    raw = raw.strip().rstrip("/")
    if not raw: return ""
    if not raw.startswith("http"): raw = "https://" + raw
    return raw + "/"

def fetch(url: str) -> tuple[str, str]:
    try:
        r = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        r.raise_for_status()
        time.sleep(DELAY_PER_REQUEST)
        return r.url, r.text
    except Exception as e:
        log.debug(f"fetch failed: {url} → {e}")
        return "", ""

def to_soup(html: str) -> BeautifulSoup:
    return BeautifulSoup(html, "lxml")

def to_txt(html: str) -> str:
    soup = to_soup(html)
    for tag in soup(["script", "style", "nav", "head"]):
        tag.decompose()
    for td in soup.find_all(["td", "th"]):
        td.insert_before(" | ")
        td.insert_after(" | ")
    return re.sub(r"\s+", " ", soup.get_text(" ")).strip()

def domain(url: str) -> str:
    return urlparse(url).netloc.lower().replace("www.", "")

def base_url(url: str) -> str:
    p = urlparse(url)
    return f"{p.scheme}://{p.netloc}"


def discover_subpages(home_url: str, home_html: str) -> list[str]:
    s = to_soup(home_html)
    base = domain(home_url)
    found = []
    for a in s.find_all("a", href=True):
        href = a["href"].strip()
        if not href or href.startswith("#") or href.startswith("mailto:"):
            continue
        abs_url = urljoin(home_url, href)
        if domain(abs_url) != base:
            continue
        path = abs_url.lower()
        for kws in SUBPAGE_KEYWORDS.values():
            if any(kw in path for kw in kws):
                if abs_url not in found:
                    found.append(abs_url)
                break
    return found[:8]


def try_direct(root: str, category: str, fetched: list) -> list[tuple[str, str]]:
    results = []
    for path in DIRECT_URL_PATTERNS.get(category, []):
        url = root.rstrip("/") + path
        if url in fetched:
            continue
        _, html = fetch(url)
        if html and len(html) > 500:
            results.append((url, html))
            if len(results) >= 2:
                break
    return results


def extract_rankings(soup: BeautifulSoup) -> list[str]:
    results = []
    seen = set()
    for tag in soup.find_all(["li", "p", "h2", "h3", "h4", "div"]):
        text = clean(tag.get_text(" "))
        if not text or len(text) < 10 or len(text) > 300:
            continue
        if any(w in text.lower() for w in NAV_SKIP):
            continue
        is_ranking = any(re.search(p, text, re.I) for p in RANKING_PATTERNS)
        if not is_ranking:
            is_ranking = any(w in text.lower() for w in RANKING_BOOST_WORDS)
        if is_ranking and text not in seen:
            seen.add(text)
            results.append(text)
            if len(results) >= 15:
                break
    return results


def extract_awards(soup: BeautifulSoup) -> list[str]:
    results = []
    seen = set()
    for tag in soup.find_all(["li", "p", "h2", "h3", "h4", "div"]):
        text = clean(tag.get_text(" "))
        if not text or len(text) < 10 or len(text) > 300:
            continue
        if any(w in text.lower() for w in NAV_SKIP):
            continue
        is_award = any(re.search(p, text, re.I) for p in AWARD_PATTERNS)
        if not is_award:
            if any(w in text.lower() for w in AWARD_BOOST_WORDS):
                if re.search(r'[A-Z][a-z]+', text):
                    is_award = True
        if is_award and text not in seen:
            seen.add(text)
            results.append(text)
            if len(results) >= 15:
                break
    return results


def scrape_rankings_awards(raw_url: str, debug: bool = False) -> tuple[list, list]:
    """Scrape rankings and awards from a university website."""
    url = normalize_url(raw_url)
    if not url:
        return [], []

    final_url, home_html = fetch(url)
    if not home_html and url.startswith("https://"):
        final_url, home_html = fetch(url.replace("https://", "http://"))
    if not home_html:
        return [], []

    root = base_url(final_url or url)
    all_soups  = [to_soup(home_html)]
    fetched    = [final_url or url]

    # Discover from homepage links
    for sub in discover_subpages(final_url or url, home_html):
        _, html = fetch(sub)
        if html:
            all_soups.append(to_soup(html))
            fetched.append(sub)

    # Try direct URL patterns
    for cat in ["rankings", "facts", "about"]:
        for sub_url, html in try_direct(root, cat, fetched):
            all_soups.append(to_soup(html))
            fetched.append(sub_url)

    all_rankings, all_awards = [], []
    seen_r, seen_a = set(), set()

    for soup in all_soups:
        for r in extract_rankings(soup):
            if r not in seen_r:
                seen_r.add(r)
                all_rankings.append(r)
        for a in extract_awards(soup):
            if a not in seen_a:
                seen_a.add(a)
                all_awards.append(a)

    if debug:
        print(f"\n  Pages fetched ({len(fetched)}):")
        for u in fetched:
            print(f"    {u}")
        print(f"  Rankings found: {len(all_rankings)}")
        for r in all_rankings:
            print(f"    ★ {r}")
        print(f"  Awards found: {len(all_awards)}")
        for a in all_awards:
            print(f"    ✦ {a}")

    return all_rankings[:15], all_awards[:15]


# ──────────────────────────────────────────────────────
#  MAIN BATCH RUNNER
# ──────────────────────────────────────────────────────
def run_batch(
    input_xlsx:    str  = "Uni_data.xlsx",
    output_xlsx:   str  = "Uni_data_filled.xlsx",
    progress_json: str  = "university_profiles.json",
    api_key:       str  = None,
):
    # ── Step 1: Load Scorecard data if API key provided ──
    scorecard_data = {}
    if api_key:
        log.info("Step 1/2: Loading College Scorecard API data...")
        scorecard_data = fetch_all_scorecard(api_key)
        log.info(f"  Scorecard ready: {len(scorecard_data):,} schools")
    else:
        log.info("No API key provided — numeric fields will be scraped (lower fill rate)")
        log.info("Get a free key at https://api.data.gov/signup")
        log.info("Then run: python university_scraper.py YOUR_KEY Uni_data.xlsx")

    # ── Step 2: Load Excel ──
    log.info(f"\nStep 2/2: Processing universities from {input_xlsx}...")
    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    # Load resume progress
    progress: dict = {}
    if os.path.exists(progress_json):
        with open(progress_json, "r", encoding="utf-8") as f:
            progress = json.load(f)
        log.info(f"  Resuming — {len(progress)} already done.")

    total   = ws.max_row - 1
    done    = 0
    skipped = 0

    for row_num in range(2, ws.max_row + 1):
        idx     = row_num - 1
        name    = clean(ws.cell(row=row_num, column=col_idx["name"]).value)
        raw_url = clean(ws.cell(row=row_num, column=col_idx["school_url"]).value)

        if not raw_url:
            skipped += 1
            continue

        # Skip if already fully processed
        if name in progress and progress[name].get("_scraped"):
            _write_row(ws, row_num, col_idx, progress[name])
            skipped += 1
            log.info(f"[{idx}/{total}] SKIP: {name}")
            continue

        log.info(f"[{idx}/{total}] → {name}")

        # Get numeric fields from Scorecard API
        numeric = scorecard_lookup(name, raw_url, scorecard_data) if scorecard_data else {}

        # Get rankings & awards from web scraping
        try:
            rankings, awards = scrape_rankings_awards(raw_url)
        except Exception as e:
            log.warning(f"  Rankings scrape failed: {e}")
            rankings, awards = [], []

        # Merge
        result = {
            "acceptance_rate":          numeric.get("acceptance_rate"),
            "in_state_tuition":         numeric.get("in_state_tuition"),
            "out_of_state_tuition":     numeric.get("out_of_state_tuition"),
            "total_cost_of_attendance": numeric.get("total_cost_of_attendance"),
            "student_size":             numeric.get("student_size"),
            "average_sat":              numeric.get("average_sat"),
            "average_act":              numeric.get("average_act"),
            "graduation_rate_4yr":      numeric.get("graduation_rate_4yr"),
            "median_earnings_10yr":     numeric.get("median_earnings_10yr"),
            "median_debt":              numeric.get("median_debt"),
            "rankings":                 rankings if rankings else None,
            "awards":                   awards   if awards   else None,
            "_scraped": True,
        }

        progress[name] = result
        _write_row(ws, row_num, col_idx, result)
        done += 1

        num_filled = sum(1 for k in FILLABLE_COLS if result.get(k) is not None)
        log.info(f"  Fields: {num_filled}/10 | Rankings: {len(rankings)} | Awards: {len(awards)}")

        # Save JSON after every university
        with open(progress_json, "w", encoding="utf-8") as f:
            json.dump(progress, f, indent=2, ensure_ascii=False)

        # Save Excel every 50
        if done % 50 == 0:
            wb.save(output_xlsx)
            log.info(f"  ✓ Checkpoint → {output_xlsx}")

    wb.save(output_xlsx)
    log.info(f"\n{'='*55}")
    log.info(f"COMPLETE  |  Processed: {done}  Skipped: {skipped}")
    log.info(f"Output    →  {output_xlsx}")
    _print_summary(progress)


def _write_row(ws, row_num, col_idx, result):
    for col in FILLABLE_COLS:
        if col in col_idx and col in result:
            ws.cell(row=row_num, column=col_idx[col]).value = result[col]


def _print_summary(progress):
    total = len(progress)
    if not total: return
    log.info(f"\n{'Field':<35} {'Found':>7} {'Fill %':>8}")
    log.info("-" * 52)
    counts = {c: sum(1 for p in progress.values() if p.get(c) is not None) for c in FILLABLE_COLS}
    for col, cnt in sorted(counts.items(), key=lambda x: -x[1]):
        log.info(f"{col:<35} {cnt:>7}  {cnt/total*100:>7.1f}%")
    r = sum(1 for p in progress.values() if p.get("rankings"))
    a = sum(1 for p in progress.values() if p.get("awards"))
    log.info(f"{'rankings':<35} {r:>7}  {r/total*100:>7.1f}%")
    log.info(f"{'awards':<35} {a:>7}  {a/total*100:>7.1f}%")


# ──────────────────────────────────────────────────────
#  ENTRY POINT
# ──────────────────────────────────────────────────────
if __name__ == "__main__":
    args = sys.argv[1:]

    # debug mode: python university_scraper.py debug https://unt.edu/ [api_key]
    if args and args[0] == "debug":
        if len(args) < 2:
            print("Usage: python university_scraper.py debug https://university.edu/ [api_key]")
            sys.exit(1)
        debug_url = args[1]
        debug_key = args[2] if len(args) > 2 else None
        print(f"\n{'='*55}")
        print(f"DEBUG: {debug_url}")

        if debug_key:
            # Quick single-school scorecard lookup
            sc = fetch_all_scorecard(debug_key)
            print(f"\n--- SCORECARD NUMERIC DATA ---")
            # Find closest match
            url_domain = domain(normalize_url(debug_url))
            matched = None
            for name, data in sc.items():
                matched = name
                break  # Just show first to verify connection
            if sc:
                print(f"  API connected ✓ ({len(sc):,} schools loaded)")
            else:
                print("  API failed ✗")

        print(f"\n--- RANKINGS & AWARDS SCRAPE ---")
        rankings, awards = scrape_rankings_awards(debug_url, debug=True)
        print(f"\nFinal: {len(rankings)} rankings, {len(awards)} awards")

    # test mode: python university_scraper.py test https://unt.edu/
    elif args and args[0] == "test":
        if len(args) < 2:
            print("Usage: python university_scraper.py test https://university.edu/")
            sys.exit(1)
        rankings, awards = scrape_rankings_awards(args[1])
        print(json.dumps({"rankings": rankings, "awards": awards}, indent=2))

    # batch mode: python university_scraper.py [api_key] [file.xlsx]
    else:
        # Figure out which args are api key vs xlsx file
        api_key    = None
        input_file = "Uni_data.xlsx"

        for arg in args:
            if arg.endswith(".xlsx"):
                input_file = arg
            elif len(arg) > 20 and not arg.endswith(".xlsx"):
                api_key = arg  # API keys are long strings

        if api_key:
            log.info(f"API key provided ✓ — will use College Scorecard for numeric data")
        else:
            log.info("No API key — only scraping rankings/awards from websites")
            log.info("For numeric data (tuition, SAT etc), get a free key: https://api.data.gov/signup")
            log.info("Then run: python university_scraper.py YOUR_KEY Uni_data.xlsx")

        run_batch(
            input_xlsx  = input_file,
            output_xlsx = input_file.replace(".xlsx", "_filled.xlsx"),
            api_key     = api_key,
        )
